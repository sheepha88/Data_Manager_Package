import numpy as np
import pandas as pd
import datetime
import warnings
import openpyxl
warnings.filterwarnings("ignore")


# 오류 메세지
class printerror(Exception):
    def __init__(self, msg) :
          self.msg = msg
    def __str__(self):
          return self.msg

class Baselineerror(printerror):
    def __init__(self,Baselinename ) :
        self.msg = ("\"{}\" do not match with baselinename of dataframe ".format(Baselinename ))

class USUBJIDerror(printerror):
    def __init__(self,USUBJID ) :
        self.msg = ("\"{}\" do not exist in USUBJID of dataframe ".format(USUBJID ))


class ADJ_PICKerror(printerror): 
    def __init__(self,USUBJID ) : 
        self.msg = ( "Subject ID : \"{}\" 조정자값이 두 평가자 값과 달라 확인요망".format(USUBJID))

class First_VISIT_Only(printerror):
    def __init__(self ) :
        self.msg = "첫 방문일을 제외한 다른 방문일 포함되어있습니다. 첫 방문일만 가능합니다."

class First_VISIT_Exclude(printerror):
    def __init__(self ) :
        self.msg = "첫 방문일이 포함되어있습니다. 첫 방문일 제외해주시기 바랍니다."



### openpyxl 사용할 때 한정 함수! ###
#해당열 number의 컬럼명 get (ex. 2열 = B)
from openpyxl.utils.cell import get_column_letter

# value값으로 인덱싱 하는 함수 
# ex) find_cell(df , "TRGIND" , -1)
# clink 는 TRGIND의 위 컬럼을 인덱싱하기 싶으면 -1로 I2 에서 I1으로 인덱싱해주는 매개변수
def find_cell(dataframe , cell_value , clink):
    # 1에서 부터 데이터 행개수 +1 까지(그래야 데이터 끝까지 범위가 지정됨)
    for x in list(range(1 , dataframe.max_row+1)):
        # 1에서 부터 데이터 열개수 +1 까지(그래야 데이터 끝까지 범위가 지정됨)
        for y in list(range(1 , dataframe.max_column+1)):
            #loop돌면서 데이터에서 cell_value값 찾으러 돌아다니고 
            if dataframe.cell(row=x, column=y).value==cell_value:
                    #찾으면 해당 열과 행-1 값 return
                    #행 -1하는 이유: USUBJID(A2)의 위의값에 SUBJECTNO(A1)이라고 지정하고 싶은 것이기 때문에 우리는 A2를 통해서 A1을 찍어줘야 한다.
                    # 따라서 A1을 도출하려면 clink에 -1을 해야한다.
                    return get_column_letter(y) + str(x+clink)
                    

# MMF에서 조정자가 Pick한 Analyst 표시하고 싶을 때 사용
# 
def ADJ_PICK_Flag(dataframe, USUBJID , Baselinename , ADJUDICATOR , Analyst_1 , Analyst_2 , Flag_col , columns):

    try:
        # 해당 대상자의 baseline만 뽑아낸 테이블
        baseline_Dataframe = dataframe[ (dataframe["USUBJID"]==USUBJID) & (dataframe["VISIT"]==Baselinename)].reset_index(drop=True)
        
        # #해당 대상자의 전체 visit 뽑아낸 테이블
        # visit_Dataframe = dataframe[ (dataframe["USUBJID"]==USUBJID)].reset_index(drop=True)
        
        # 조정자의 컬럼에 해당하는 series값
        ADJ_series = baseline_Dataframe.loc[list(baseline_Dataframe[baseline_Dataframe["READER"]==ADJUDICATOR].index)[0] , columns]
        # Anaslyst#1의 컬럼에 해당하는 series값
        Analyst_1_series = baseline_Dataframe.loc[list(baseline_Dataframe[baseline_Dataframe["READER"]==Analyst_1].index)[0] , columns]
        # Anaslyst#2의 컬럼에 해당하는 series값
        Analyst_2_series = baseline_Dataframe.loc[list(baseline_Dataframe[baseline_Dataframe["READER"]==Analyst_2].index)[0] , columns]


        # 만약 조정자의 컬럼에 해당하는 series값이 모두 nan이라면(조정자가 아직 판독하지 않은 경우라면) 제외해라
        if not all([pd.isnull(x) for x in ADJ_series.values.tolist()]):

            # 조정자의 columns값과 Analyst#1의 columns값이 baseline에서 같다면 analyst#1을 선언
            if ADJ_series.equals(Analyst_1_series):
                    ADJ_Pick_Analayst = Analyst_1
                    
            # 조정자의 columns값과 Analyst#1의 columns값이 baseline에서 같다면 analyst#2을 선언
            elif ADJ_series.equals( Analyst_2_series):
                    ADJ_Pick_Analayst = Analyst_2
            
            
            # 헤당 대상자 , 조정자가 pick한 Reader를 Flag_col에 표시 , Y , N
            dataframe.loc[ (dataframe["USUBJID"]==USUBJID) & (dataframe["READER"]==ADJ_Pick_Analayst) , Flag_col] = "O" 
            dataframe.loc[ (dataframe["USUBJID"]==USUBJID) & (dataframe["READER"]!=ADJ_Pick_Analayst) , Flag_col] = np.nan 

            

    except:
        #error 출력
        #basline이 dataframe에 없는 경우
        if Baselinename not in list(dataframe["VISIT"].unique()):
            raise Baselineerror(Baselinename)

        #대상자가 dataframe에 없는 경우
        if USUBJID not in list(dataframe["USUBJID"].unique()):
            raise USUBJIDerror(USUBJID)

        #조정자 값이 두 평가자 값과 달라 확인이 필요한 경우
        if not ADJ_series.equals(Analyst_1_series):
            if not ADJ_series.equals(Analyst_2_series):
                raise ADJ_PICKerror(USUBJID)



# baeline에서 PCBSLD , PCNSLD , SUMBLD , SUMNLD 등은 np.nan값으로 바꿔주는 함수
# ex)makevalue(df , "SCRN_CT", "PCBSLD" , np.nan)
# baelineNAME : baeline이름(bl or SCRN_CT,,,) , colname : 컬럼이름 , value: 변경 후 값 (여기서는 np.nan)
# dataframe copy를 썼기 때문에 선언해줘야 한다 ex)df.loc[0,"PCBSLD"] = makevalue(df ,0 , "SCRN_CT", "PCBSLD" , 3)

def makevalue(dataframe, baselineNAME , colname , value):
    
    new_dataframe = dataframe.copy(deep = True)
    
    for i in range(len(dataframe)):

        if new_dataframe.loc[i , "VISIT"]==baselineNAME:
            new_dataframe.loc[i , colname] = value
    
    return new_dataframe





# map develop 함수 -> dictionary에 없는 값은 원래의 값을 출력
# ex) map_dict(df , "LAGRADE",LAGRADE_dict ).unique()
def map_dict(dataframe, col , dict_name):
    func = lambda x : dict_name.get(x,x)
    dataframe_new = dataframe[col].map(func , na_action = None)
    
    return dataframe_new


    